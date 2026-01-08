from evdev import InputDevice, categorize, ecodes
from smbus2 import SMBus
import time
from lcd_driver import LCD
import board
import adafruit_dht
from gpiozero import LED
from gpiozero import Button
from gpiozero import Buzzer
from line_env.send_line_message import send_name
from signal import pause
import subprocess
import sys
import os
import json
import threading
from openpyxl import Workbook, load_workbook

#excel
EXCEL_FILE = "door_log.xlsx"
MAX_RECORDS = 10
def log_door_event(time_str, method, identity, result):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["時間", "方式", "身份", "狀態"])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    # 新增紀錄
    ws.append([time_str, method, identity, result])
    # 如果超過 10 筆，刪除最舊的
    while ws.max_row > MAX_RECORDS + 1: 
        ws.delete_rows(2)
    wb.save(EXCEL_FILE)
#LED
g_led=LED(21)
r_led=LED(20)
#蜂鳴器
buzzer = Buzzer(23)
#flask
def start_flask():
    flask_path = os.path.join("line_env", "QR.py")

    print("啟動 Flask LINE Bot...")
    subprocess.Popen(
        [sys.executable, flask_path],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )
start_flask()
#QR
QR_REQUEST_FILE = "qr_verified.json"
last_qr_time = 0
def qr_watch_loop():
    while True:
        check_qr_request_and_open()
        time.sleep(0.5)

def check_qr_request_and_open():
    if not os.path.exists(QR_REQUEST_FILE):
        return

    try:
        with open(QR_REQUEST_FILE, "r") as f:
            content = f.read().strip()
            if not content:
                return  # 檔案是空的，直接跳過

            data = json.loads(content)

    except json.JSONDecodeError:
        # 正在寫入中，下一輪再讀
        return

    except Exception as e:
        print("讀取 QR 檔案錯誤：", e)
        return

    # 讀JSON
    if data.get("type") == "QR_VERIFIED":
        print("收到 QR 驗證成功")
        print("Token:", data["token"])
        print("Time:", now_time())
        lcd.message("QR SUCCESS", 2)
        open_door_by_qr()
        os.remove(QR_REQUEST_FILE)
        log_door_event(now_time(),"QR",data["token"],"SUCCESS")
        time.sleep(2)
        lcd.message("Scan your card", 2)

#door
def open_door_by_qr():
    print("開門成功")
def door_opened():
    global door_open_time, door_warning_sent
    door_open_time=time.time()
    print("時間：",now_time())
    door_warning_sent=False
    print("門被打開")
    buzzer.on()
    time.sleep(0.1)
    buzzer.off()
   
def door_closed():
    global door_open_time, door_warning_sent
    door_open_time=None
    print("時間：",now_time())
    door_warning_sent=False
    print("門已關閉")
    time.sleep(0.1)
    buzzer.off() 
#門磁感應
door_sensor = Button(17)
door_open_time = None
door_warning_sent =False
door_sensor.when_pressed=door_opened
door_sensor.when_released=door_closed
def door_warn():
    global door_warning_sent
    while True:
        if door_open_time is not None:
            if (time.time() - door_open_time) >= 10 and not door_warning_sent:
                print("門已開啟超過 10 秒")
                send_name("警告：門已開啟超過 10 秒未關閉")
                log_door_event(now_time(),"門磁","開啟超過 10 秒","")
                door_warning_sent = True
        buzzer.on()
        time.sleep(0.5)
        buzzer.off()

#LCD
lcd = LCD(2, 0x27, True)
def retime():
    lcd.message(now_time(),1)
    time.sleep(1)
#time
def now_time():
    return time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(time.time() + 8 * 3600))

#R20C‑USB 裝置
device = InputDevice('/dev/input/event5')
print("使用裝置：", device)

#授權UID
AUTHORIZED_UIDS = { "0084703874":"a","0442131957":"b"}


#主程式
lcd.message(str(now_time()),1)
lcd.message("Scan your card", 2)

uid_buffer = ""
last_time = 0

threading.Thread(target=retime, daemon=True).start()
threading.Thread(target=door_warn, daemon=True).start()
threading.Thread(target=qr_watch_loop, daemon=True).start()
try:
    current_time = time.time()
    check_qr_request_and_open()
    for event in device.read_loop():

        current_time = time.time()

        if current_time - last_time >= 60:
            lcd.message(str(now_time()), 1)
            last_time = current_time

        if event.type == ecodes.EV_KEY:
            keyevent = categorize(event)
            if keyevent.keystate == keyevent.key_down:
                key = keyevent.keycode

                if key == 'KEY_ENTER':
                    print("UID:", uid_buffer)

                    if uid_buffer in AUTHORIZED_UIDS:
                        lcd.message("SUCCESS", 2)
                        g_led.on()
                        print("授權成功")
                        print("身份：",AUTHORIZED_UIDS[uid_buffer])
                        print("時間：",now_time())
                        send_name(f"門禁解鎖\n身份：{AUTHORIZED_UIDS[uid_buffer]}\n時間：{now_time()}")
                        open_door_by_qr()
                        log_door_event(now_time(),"ID卡",AUTHORIZED_UIDS[uid_buffer],"SUCCESS")
                    else:
                        lcd.message("FAIL", 2)
                        r_led.on()
                        print("授權失敗")
                        log_door_event(now_time(),"ID卡","","FAIL")

                    time.sleep(2)
                    lcd.message(now_time(),1)
                    lcd.message("Scan your card", 2)
                    uid_buffer = ""
                    g_led.off()
                    r_led.off()

                elif key.startswith('KEY_'):
                    char = key.replace('KEY_', '')
                    if len(char) == 1:
                        uid_buffer += char
except KeyboardInterrupt :
    print("程式結束")
