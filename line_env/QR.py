#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import time
import uuid
import json
import logging

from dotenv import load_dotenv
from flask import Flask, request, abort
from openpyxl import load_workbook
from linebot.models import FileMessage
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import (
    MessageEvent,
    TextMessage,
    TextSendMessage,
    ImageSendMessage,
)

# ------------------ 初始設定 ------------------ #

load_dotenv()

CHANNEL_ACCESS_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
CHANNEL_SECRET = os.getenv("LINE_CHANNEL_SECRET")

if not CHANNEL_ACCESS_TOKEN or not CHANNEL_SECRET:
    print("請先在 .env 中設定 LINE_CHANNEL_ACCESS_TOKEN 和 LINE_CHANNEL_SECRET")
    sys.exit(1)

line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(CHANNEL_SECRET)

logging.basicConfig(level=logging.INFO)

app = Flask(__name__)

# ------------------ QR / IPC 設定 ------------------ #

VALID_QR_TOKENS = {}          # token : timestamp
QR_EXPIRE_SECONDS = 60

QR_VERIFY_FILE = "qr_verified.json"   # 給主程式讀

# ------------------ Flask 路由 ------------------ #

@app.route("/", methods=["GET"])
def index():
    return "LINE Bot is running.", 200


@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers.get("X-Line-Signature", "")
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return "OK"


@app.route("/open", methods=["GET"])
def open_by_qr():
    token = request.args.get("token")
    now = time.time()

    if not token or token not in VALID_QR_TOKENS:
        return "Invalid QR Code", 403

    if now - VALID_QR_TOKENS[token] > QR_EXPIRE_SECONDS:
        del VALID_QR_TOKENS[token]
        return "QR Code Expired", 403

    # 一次性使用
    del VALID_QR_TOKENS[token]

    # 🔔 通知主程式（QR 驗證成功）
    data = {
        "type": "QR_VERIFIED",
        "token": token,
        "time": now
    }

    with open(QR_VERIFY_FILE, "w") as f:
        json.dump(data, f)

    print("QR 驗證成功，已通知主程式")

    return "QR 驗證成功", 200
from flask import send_file

#讀excel
EXCEL_FILE = "door_log.xlsx"

def read_excel_to_text(max_rows=10):
    """讀取 Excel 檔案，回傳文字列表"""
    if not os.path.exists(EXCEL_FILE):
        return "目前沒有門禁紀錄。"

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    # 抓標題
    rows = list(ws.iter_rows(values_only=True))
    text = "門禁紀錄（最新10筆）\n"
    headers = rows[0]
    text += " | ".join(headers) + "\n"
    text += "-"*40 + "\n"

    # 取最新 max_rows 筆
    for row in rows[-max_rows:]:
        text += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"

    return text
# ------------------ LINE 訊息處理 ------------------ #

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event: MessageEvent):
    user_text = event.message.text.strip()


    if user_text == "給我QR":
        token = str(uuid.uuid4())
        VALID_QR_TOKENS[token] = time.time()

        # ⚠️ 若之後用 ngrok，改成 ngrok 網址
        
        pi_ip = "192.168.1.37" #ip a
        qr_target = f"http://{pi_ip}:5000/open?token={token}"
        # 使用線上 QR 產生服務
        qr_image_url = (
            "https://api.qrserver.com/v1/create-qr-code/"
            f"?size=300x300&data={qr_target}"
        )

        line_bot_api.reply_message(
            event.reply_token,
            ImageSendMessage(
                original_content_url=qr_image_url,
                preview_image_url=qr_image_url
            )
        )
    elif user_text=="門禁紀錄":
        excel_text = read_excel_to_text(max_rows=10)
        line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text=excel_text)
        )
    else:
        reply_text = (
            "可用指令：\n"
            "給我QR（產生開門 QR Code）"
            "門禁紀錄"
        )

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=reply_text)
        )

# ------------------ 主程式入口 ------------------ #

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
